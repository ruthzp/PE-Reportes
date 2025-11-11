# funciones_op1.py
# ============================================================
# Módulo de funciones para la hoja OP1
# ============================================================

import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.workbook.properties import CalcProperties
import streamlit as st


# ============================================================
# FUNCIONES AUXILIARES GENERALES
# ============================================================

def habilitar_recalculo(wb):
    """Activa recálculo completo al abrir Excel."""
    try:
        wb.calculation_properties.fullCalcOnLoad = True
    except Exception:
        wb.calculation_properties = CalcProperties(fullCalcOnLoad=True)


def cargar_excel_con_encabezado_correcto(file):
    """Detecta la fila donde aparece 'Sede Operativa' y la usa como encabezado."""
    df_raw = pd.read_excel(file, header=None)
    header_row = None
    for i in range(len(df_raw)):
        fila = df_raw.iloc[i].astype(str).str.lower()
        if fila.str.contains("sede operativa").any():
            header_row = i
            break
    if header_row is None:
        raise ValueError("❌ No se encontró la fila con 'Sede Operativa' en el archivo.")
    df = pd.read_excel(file, header=header_row)
    df.columns = df.columns.str.strip()
    return df


# ============================================================
# FUNCIÓN: GENERAR HOJA OP1
# ============================================================

def generar_op1(base, asc_fa, asc_inst, nom_inst):
    """Genera la hoja OP1 completa según las reglas definidas."""
    st.info("Procesando hoja OP1...")

    try:
        # 1️⃣ Cargar datos con detección de encabezado
        asc_fa_df = cargar_excel_con_encabezado_correcto(asc_fa)
        asc_inst_df = cargar_excel_con_encabezado_correcto(asc_inst)
        nom_inst_df = cargar_excel_con_encabezado_correcto(nom_inst)

        # 2️⃣ Abrir archivo base
        wb = load_workbook(base)
        if "OP1" not in wb.sheetnames:
            raise ValueError("❌ No se encontró la hoja 'OP1' en el archivo base.")

        # 🔹 Dejar solo la hoja OP1
        for nombre in wb.sheetnames.copy():
            if nombre != "OP1":
                del wb[nombre]

        ws = wb["OP1"]

        # 3️⃣ Actualizar hoja
        actualizar_OP1(ws, asc_fa_df, asc_inst_df, nom_inst_df)

        # 4️⃣ Guardar salida y mantener en memoria
        habilitar_recalculo(wb)
        out = BytesIO()
        wb.save(out)
        out.seek(0)

        # ✅ Mantener archivo en sesión para que no desaparezca el botón
        st.session_state["op1_generada"] = out

        st.success("✅ Hoja OP1 generada correctamente.")

    except Exception as e:
        st.error(f"❌ Error al generar OP1: {e}")


# ============================================================
# LÓGICA PRINCIPAL: ACTUALIZAR OP1
# ============================================================

def actualizar_OP1(ws, asc_fa_df, asc_inst_df, nom_inst_df):
    """Actualiza todos los valores y fórmulas de la hoja OP1."""

    asc_fa_sede, asc_fa_local = "Sede Operativa", "Local"
    asc_inst_sede, asc_inst_local = "Sede Operativa", "Local"
    nom_inst_sede, nom_inst_local = "Sede Operativa", "Local"
    col_inventario = "Inventario en campo"

    # === Iterar sobre todas las filas de OP1 ===
    for r in range(2, ws.max_row + 1):
        sede = str(ws[f"B{r}"].value or "").strip()
        local = str(ws[f"C{r}"].value or "").strip()
        if not sede or not local:
            continue

        # =====================================================
        # BLOQUE ASC - INSTRUMENTOS
        # =====================================================
        asc_c = asc_inst_df.loc[
            (asc_inst_df[asc_inst_sede].astype(str).str.strip() == sede)
            & (asc_inst_df[asc_inst_local].astype(str).str.strip() == local)
            & (asc_inst_df["Tipo"].str.contains("CUADERNILLO DE CONOCIMIENTOS PEDAGÓGICOS", case=False, na=False)),
            col_inventario,
        ].sum()

        asc_f = asc_inst_df.loc[
            (asc_inst_df[asc_inst_sede].astype(str).str.strip() == sede)
            & (asc_inst_df[asc_inst_local].astype(str).str.strip() == local)
            & (asc_inst_df["Tipo"].str.contains("FICHA DE RESPUESTA", case=False, na=False)),
            col_inventario,
        ].sum()

        ws[f"M{r}"].value = asc_c
        ws[f"N{r}"].value = asc_f
        ws[f"O{r}"].value = f"=G{r}-M{r}"
        ws[f"P{r}"].value = f"=H{r}-N{r}"
        ws[f"Q{r}"].value = f"=IF(G{r}=0,1,M{r}/G{r})"
        ws[f"R{r}"].value = f"=IF(H{r}=0,1,N{r}/H{r})"

        # =====================================================
        # BLOQUE NOM - INSTRUMENTOS
        # =====================================================
        nom_local = nom_inst_df[
            (nom_inst_df[nom_inst_sede].astype(str).str.strip() == sede)
            & (nom_inst_df[nom_inst_local].astype(str).str.strip() == local)
        ]

        if not nom_local.empty:
            nom_c = nom_local.loc[
                nom_local["Tipo"].str.contains(
                    "CUADERNILLO DE CONOCIMIENTOS PEDAGÓGICOS|CUADERNILLO DE HABILIDADES GENERALES",
                    case=False,
                    na=False,
                ),
                col_inventario,
            ].sum()
            nom_f = nom_local.loc[
                nom_local["Tipo"].str.contains("FICHA DE RESPUESTA", case=False, na=False),
                col_inventario,
            ].sum()
        else:
            nom_c = nom_f = 0

        ws[f"S{r}"].value = nom_c
        ws[f"T{r}"].value = nom_f
        ws[f"U{r}"].value = f"=I{r}-S{r}"
        ws[f"V{r}"].value = f"=J{r}-T{r}"
        ws[f"W{r}"].value = f"=IF(I{r}=0,1,S{r}/I{r})"
        ws[f"X{r}"].value = f"=IF(J{r}=0,1,T{r}/J{r})"

        # =====================================================
        # BLOQUE ASC - FA (Formatos Auxiliares)
        # =====================================================
        def sumar_tipo(tipo):
            return asc_fa_df.loc[
                (asc_fa_df[asc_fa_sede].astype(str).str.strip() == sede)
                & (asc_fa_df[asc_fa_local].astype(str).str.strip() == local)
                & (asc_fa_df["Tipo"].str.contains(tipo, case=False, na=False)),
                col_inventario,
            ].sum()

        tipos = {
            "AN": "ACTA DE RECEPCIÓN/DEVOLUCIÓN",
            "AP": "ACTA DE APLICACIÓN DEL AULA",
            "AR": "LISTA DE ASISTENCIA",
            "AT": "LISTA DE RETIRO DE CUADERNILLOS",
            "AV": "ACTA DE RESPUESTA A OBSERVACIONES DEL DOCENTE",
            "AX": "REGISTRO DE ENTREGA INSTRUMENTOS ADICIONALES",
            "AZ": "ACTA DE INCIDENCIAS DEL CAE",
            "BB": "ACTA DE INCUMPLIMIENTO DE PROCEDIMIENTOS",
            "BD": "ACTA DE INCIDENCIAS DE SALUD",
            "BF": "ACTA DE INCIDENCIAS DEL LOCAL DE EVALUACIÓN",
            "BH": "ACTA FISCAL",
            "BJ": "SOBRES",
        }

        for col, tipo in tipos.items():
            ws[f"{col}{r}"].value = sumar_tipo(tipo)

        # =====================================================
        # FÓRMULAS DE PORCENTAJES Y VALIDACIONES
        # =====================================================
        ws[f"AO{r}"].value = f"=AN{r}/AB{r}"
        ws[f"AQ{r}"].value = f"=AP{r}/AC{r}"
        ws[f"AS{r}"].value = f"=AR{r}/AD{r}"
        ws[f"AU{r}"].value = f"=AT{r}/AE{r}"
        ws[f"AW{r}"].value = f"=AV{r}/AF{r}"
        ws[f"AY{r}"].value = f"=AX{r}/AG{r}"
        ws[f"BA{r}"].value = f"=AZ{r}/AH{r}"
        ws[f"BC{r}"].value = f'=IF(BB{r}=AI{r},"OK","ERR")'
        ws[f"BE{r}"].value = f"=BD{r}/AJ{r}"
        ws[f"BG{r}"].value = f"=BF{r}/AK{r}"
        ws[f"BI{r}"].value = f'=IF(BH{r}=AL{r},"OK","ERR")'
        ws[f"BK{r}"].value = f"=BJ{r}/AM{r}"

    return ws
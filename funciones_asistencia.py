import pandas as pd
import streamlit as st
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import CellIsRule

def detectar_columna_sede(df):
    for c in df.columns:
        if any(x in str(c).lower() for x in ["sede", "operativa", "evaluación", "aplicación"]):
            return c
    raise ValueError("❌ No se encontró columna de sede válida.")

def cargar_postulantes(file):
    df = pd.read_excel(file, header=None)
    i = df.index[df.iloc[:, 0].astype(str).str.upper().eq("N")].tolist()
    if not i:
        raise ValueError("❌ No se encontró cabecera con 'N'.")
    df.columns = df.iloc[i[0]]
    df = df.iloc[i[0] + 1:].reset_index(drop=True)
    sede_col = detectar_columna_sede(df)
    df = df.rename(columns={sede_col: "Sede"})
    for c in ["Postulantes", "Asistencia al Local", "Asistencia en Aula", "Casos de inconsistencia"]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    return df.groupby("Sede", as_index=False).sum()

def habilitar_recalculo(wb):
    from openpyxl.workbook.properties import CalcProperties
    try:
        if not hasattr(wb, "calculation_properties"):
            wb.create_sheet("_tmp_")
            wb.remove(wb["_tmp_"])
        wb.calculation_properties.fullCalcOnLoad = True
    except Exception:
        wb.calculation_properties = CalcProperties(fullCalcOnLoad=True)

def generar_asistencia(base, asc, nom, acc):
    st.info("Procesando hoja ASISTENCIA...")
    try:
        asc_df, nom_df, acc_df = map(cargar_postulantes, [asc, nom, acc])
        asc_d, nom_d, acc_d = (d.set_index("Sede").to_dict("index") for d in [asc_df, nom_df, acc_df])

        wb = load_workbook(base)
        if "ASISTENCIA" not in wb.sheetnames:
            raise ValueError("❌ No existe la hoja 'ASISTENCIA' en el archivo base.")

        # 🔹 Dejar solo la hoja ASISTENCIA en el archivo
        for nombre in wb.sheetnames.copy():
            if nombre != "ASISTENCIA":
                del wb[nombre]

        ws = wb["ASISTENCIA"]

        rojo, verde = PatternFill("solid", fgColor="FFC7CE"), PatternFill("solid", fgColor="C6EFCE")

        for r in range(2, ws.max_row + 1):
            sede = str(ws[f"B{r}"].value or "").strip()
            if not sede:
                continue

            get = lambda d, k: d.get(sede, {}).get(k, 0)
            asc_p, asc_l, asc_a, asc_i = [get(asc_d, k) for k in ["Postulantes", "Asistencia al Local", "Asistencia en Aula", "Casos de inconsistencia"]]
            nom_p, nom_l, nom_a, nom_i = [get(nom_d, k) for k in ["Postulantes", "Asistencia al Local", "Asistencia en Aula", "Casos de inconsistencia"]]
            acc_p, acc_l, acc_a, acc_i = [get(acc_d, k) for k in ["Postulantes", "Asistencia al Local", "Asistencia en Aula", "Casos de inconsistencia"]]

            for c, v in zip("FGHI", [asc_p, asc_l, asc_a, asc_i]): ws[f"{c}{r}"].value = v
            for c, v in zip("JKLM", [nom_p, nom_l, nom_a, nom_i]): ws[f"{c}{r}"].value = v
            for c, v in zip("STUV", [acc_p, acc_l, acc_a, acc_i]): ws[f"{c}{r}"].value = v

            ws[f"N{r}"].value = f"=F{r}+J{r}"
            ws[f"O{r}"].value = f"=G{r}+K{r}"
            ws[f"P{r}"].value = f"=H{r}+L{r}"
            ws[f"Q{r}"].value = f"=I{r}+M{r}"
            ws[f"R{r}"].value = f'=IF($D{r}=$Q{r},"OK","ERR")'
            ws[f"W{r}"].value = f"=S{r}"
            ws[f"X{r}"].value = f"=T{r}"
            ws[f"Y{r}"].value = f"=U{r}"
            ws[f"Z{r}"].value = f"=V{r}"
            ws[f"AA{r}"].value = f'=IF($Z{r}=0,"OK","ERR")'

        for c in ["R", "AA"]:
            ws.conditional_formatting.add(f"{c}2:{c}{ws.max_row}", CellIsRule("equal", ['"ERR"'], fill=rojo))
            ws.conditional_formatting.add(f"{c}2:{c}{ws.max_row}", CellIsRule("equal", ['"OK"'], fill=verde))

        habilitar_recalculo(wb)

        out = BytesIO()
        wb.save(out)
        out.seek(0)
        st.session_state["asistencia_generada"] = out

        st.success("✅ Hoja ASISTENCIA generada correctamente. Puedes descargarla abajo ⬇️")

    except Exception as e:
        st.error(f"❌ Error al generar hoja ASISTENCIA: {e}")
# funciones_op2.py
# ============================================================
# Módulo de funciones para la hoja OP2 (ACCESO)
# ============================================================

import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.workbook.properties import CalcProperties
import unicodedata
import streamlit as st


# ============================================================
# FUNCIONES AUXILIARES
# ============================================================

def habilitar_recalculo(wb):
    """Activa recálculo completo al abrir Excel."""
    try:
        wb.calculation_properties.fullCalcOnLoad = True
    except Exception:
        wb.calculation_properties = CalcProperties(fullCalcOnLoad=True)


def normalizar_texto(valor):
    """Convierte texto a minúsculas, sin tildes ni guiones especiales."""
    if pd.isna(valor):
        return ""
    return (
        unicodedata.normalize('NFKD', str(valor).strip().lower())
        .encode('ascii', 'ignore')
        .decode('utf-8')
        .replace('–', '-')
        .replace('—', '-')
        .replace('  ', ' ')
    )


def cargar_excel_con_encabezado_correcto(file):
    """Detecta la fila donde aparece 'Sede Operativa' y la usa como encabezado."""
    df_raw = pd.read_excel(file, header=None)
    header_row = None
    for i in range(len(df_raw)):
        fila = df_raw.iloc[i].astype(str).str.lower()
        if fila.str.contains("sede operativa").any():
            header_row = i
            break
    if header_row is None:
        raise ValueError("❌ No se encontró la fila con 'Sede Operativa' en el archivo.")
    df = pd.read_excel(file, header=header_row)
    df.columns = df.columns.str.strip()
    return df


# ============================================================
# FUNCIÓN PRINCIPAL
# ============================================================

def generar_op2(base, acc_fa, acc_inst):
    """
    Genera la hoja OP2 en el archivo base (PE3 - Reporte.xlsx)
    usando:
      - ACC - INSTRUMENTOS.xlsx
      - ACC - FA.xlsx
    """
    st.info("⚙️ Procesando hoja OP2...")

    try:
        # 1️⃣ Cargar los datos ACC
        acc_fa_df = cargar_excel_con_encabezado_correcto(acc_fa)
        acc_inst_df = cargar_excel_con_encabezado_correcto(acc_inst)

        # Normalizar textos en DataFrames
        for df in [acc_fa_df, acc_inst_df]:
            for col in ["Sede Operativa", "Local", "Tipo"]:
                df[col] = df[col].astype(str).map(normalizar_texto)

        # 2️⃣ Abrir plantilla base
        wb = load_workbook(base)
        if "OP2" not in wb.sheetnames:
            raise ValueError("❌ No se encontró la hoja 'OP2' en el archivo base.")

        # 🔹 Dejar solo la hoja OP2
        for nombre in wb.sheetnames.copy():
            if nombre != "OP2":
                del wb[nombre]

        ws = wb["OP2"]

        # 3️⃣ Procesar hoja
        actualizar_OP2(ws, acc_fa_df, acc_inst_df)

        # 4️⃣ Forzar recálculo al abrir (compatible con todas las versiones de openpyxl)
        try:
            wb.calculation_properties.fullCalcOnLoad = True
            wb.calculation_properties.calcMode = "auto"
            wb.calculation_properties.calcId = 0
        except Exception:
            wb.calculation_properties = CalcProperties(fullCalcOnLoad=True)

        ws.calculationId = None
        wb.active = wb.sheetnames.index("OP2")

        # 5️⃣ Guardar resultado en memoria
        habilitar_recalculo(wb)
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        st.session_state["op2_generada"] = out

        st.success("✅ Hoja OP2 generada correctamente.")

    except Exception as e:
        st.error(f"❌ Error al generar OP2: {e}")


# ============================================================
# FUNCIÓN PRINCIPAL DE CÁLCULO
# ============================================================

def actualizar_OP2(ws, acc_fa_df, acc_inst_df):
    """
    Actualiza fila por fila la hoja OP2:

    - A–H: se mantienen.
    - I–J: se llenan desde ACC-INSTRUMENTOS.
    - K–N: fórmulas automáticas.
    - O–AC: se conservan.
    - AD–AZ: datos desde ACC-FA.
    - AE–BA: fórmulas automáticas.
    """

    sede_col = "Sede Operativa"
    local_col = "Local"
    tipo_col = "Tipo"
    inv_col = "Inventario en campo"

    # Iterar filas
    for r in range(2, ws.max_row + 1):
        sede = normalizar_texto(ws[f"B{r}"].value)
        local = normalizar_texto(ws[f"C{r}"].value)
        if not sede or not local:
            continue

        # ----------------------------------------------------
        # 1️⃣ ACC - INSTRUMENTOS → columnas I, J
        # ----------------------------------------------------
        df_inst = acc_inst_df[
            (acc_inst_df[sede_col] == sede) &
            (acc_inst_df[local_col] == local)
        ]

        acc_c = df_inst.loc[
            df_inst[tipo_col].str.contains("cuadernillo de conocimientos pedagog", case=False, na=False),
            inv_col
        ].sum()

        acc_f = df_inst.loc[
            df_inst[tipo_col].str.contains("ficha de respuesta", case=False, na=False),
            inv_col
        ].sum()

        ws[f"I{r}"].value = acc_c if pd.notna(acc_c) else 0
        ws[f"J{r}"].value = acc_f if pd.notna(acc_f) else 0

        # ----------------------------------------------------
        # 2️⃣ Fórmulas K–N (con IF en inglés)
        # ----------------------------------------------------
        ws[f"K{r}"].value = f"=E{r}-I{r}"
        ws[f"L{r}"].value = f"=F{r}-J{r}"
        ws[f"M{r}"].value = f"=IF(E{r}=0,1,I{r}/E{r})"
        ws[f"N{r}"].value = f"=IF(F{r}=0,1,J{r}/F{r})"

        # ----------------------------------------------------
        # 3️⃣ ACC - FA → columnas AD, AF, AH, AJ, AL, AN, AP, AR, AT, AV, AX, AZ
        # ----------------------------------------------------
        df_fa = acc_fa_df[
            (acc_fa_df[sede_col] == sede) &
            (acc_fa_df[local_col] == local)
        ]

        def sumar_tipo(patrones):
            if df_fa.empty:
                return 0
            mask = df_fa[tipo_col].str.contains("|".join(patrones), case=False, na=False)
            return df_fa.loc[mask, inv_col].sum()

        tipos = {
            "AD": ["ACTA DE RECEPCION/DEVOLUCION", "ACTA DE RECEPCIÓN/DEVOLUCIÓN"],
            "AF": ["ACTA DE APLICACION DEL AULA", "ACTA DE APLICACIÓN DEL AULA"],
            "AH": ["LISTA DE ASISTENCIA"],
            "AJ": ["LISTA DE RETIRO DE CUADERNILLOS"],
            "AL": ["ACTA DE RESPUESTA A OBSERVACIONES DEL DOCENTE"],
            "AN": ["REGISTRO DE ENTREGA INSTRUMENTOS ADICIONALES"],
            "AP": ["ACTA DE INCIDENCIAS DEL CAE"],
            "AR": ["ACTA DE INCUMPLIMIENTO DE PROCEDIMIENTOS"],
            "AT": ["ACTA DE INCIDENCIAS DE SALUD"],
            "AV": ["ACTA DE INCIDENCIAS DEL LOCAL DE EVALUACION", "ACTA DE INCIDENCIAS DEL LOCAL DE EVALUACIÓN"],
            "AX": ["ACTA FISCAL"],
            "AZ": ["SOBRES", "SOBRE"],
        }

        for col, pats in tipos.items():
            val = sumar_tipo(pats)
            ws[f"{col}{r}"].value = val if pd.notna(val) else 0

        # ----------------------------------------------------
        # 4️⃣ Fórmulas AE–BA (todas con IF en inglés)
        # ----------------------------------------------------
        formulas = {
            "AE": f"=AD{r}/R{r}",
            "AG": f"=AF{r}/S{r}",
            "AI": f"=AH{r}/T{r}",
            "AK": f"=AJ{r}/U{r}",
            "AM": f"=AL{r}/V{r}",
            "AO": f"=AN{r}/W{r}",
            "AQ": f"=AP{r}/X{r}",
            "AS": f'=IF(AR{r}=Y{r},"OK","ERR")',
            "AU": f"=AT{r}/Z{r}",
            "AW": f"=AV{r}/AA{r}",
            "AY": f'=IF(AX{r}=AB{r},"OK","ERR")',
            "BA": f"='OP2'!$AZ{r}/'OP2'!$AC{r}",
        }

        for col, formula in formulas.items():
            ws[f"{col}{r}"].value = formula

    return ws